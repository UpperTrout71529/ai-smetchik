import streamlit as st
import requests
import json
import base64
import pandas as pd
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="ИИ Сметчик материалов", page_icon="📊", layout="centered")

st.title("📊 Автоматизированный ИИ-расчет смет (Профессиональная версия)")
st.write("Загрузите проектный PDF со схемами или спецификациями. Система извлечет данные через ИИ по жесткой схеме и сформирует премиальный Excel.")

# Выбор режима работы шлюза
api_provider = st.radio(
    "🌐 Выберите способ подключения к ИИ:",
    ["ProxyAPI (Для РФ, без VPN, стабильно)", "Google AI Studio Напрямую (Нужен VPN/Прокси)"]
)

if api_provider.startswith("ProxyAPI"):
    gemini_api_key = st.text_input("🔑 Введите ваш ProxyAPI Ключ (pa-...):", type="password")
else:
    gemini_api_key = st.text_input("🔑 Введите ваш Gemini API Key:", type="password")

uploaded_file = st.file_uploader("Выберите PDF-файл проекта (чертежи или спецификации)", type=["pdf"])

if uploaded_file is not None:
    st.success(f"Файл '{uploaded_file.name}' загружен!")
    
    if st.button("🚀 Запустить профессиональный расчет"):
        if not gemini_api_key:
            st.error("Ошибка: Пожалуйста, введите ваш API-ключ.")
            st.stop()
        
        with st.status("Выполнение процессов...", expanded=True) as status:
            
            # --- ШАГ 1: Кодирование файла ---
            status.update(label="Шаг 1: Оптимизация и кодирование чертежей...", state="running")
            bytes_data = uploaded_file.getvalue()
            pdf_base64 = base64.b64encode(bytes_data).decode("utf-8")
            
            # --- ШАГ 2: Запрос к Gemini со Structured Outputs ---
            status.update(label="Шаг 2: Анализ схем и извлечение спецификации ИИ...", state="running")
            
            # ИСПОЛЬЗУЕМ СТРОГОЕ, СТАБИЛЬНОЕ ИМЯ МОДЕЛИ
            model_name = "gemini-1.5-flash"
            
            if api_provider.startswith("ProxyAPI"):
                # Собираем URL динамически с подстановкой model_name
                gemini_url = f"https://api.proxyapi.ru/google/v1beta/models/{model_name}:generateContent"
                headers = {
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {gemini_api_key}"
                }
            else:
                gemini_url = f"https://generativelanguage.googleapis.com/v1beta/models/{model_name}:generateContent"
                headers = {
                    "Content-Type": "application/json",
                    "X-goog-api-key": gemini_api_key
                }
            
            prompt = "Ты — ведущий инженер-сметчик. Внимательно изучи переданный документ (спецификации или однолинейные схемы). Найди все материалы, кабели, щиты и оборудование. Сформируй строгий список объектов по указанной JSON-схеме. Пропускай текстовые описания проекта, собирай только номенклатуру для закупки."

            payload = {
                "contents": [{
                    "parts": [
                        {"text": prompt},
                        {"inlineData": {"mimeType": "application/pdf", "data": pdf_base64}}
                    ]
                }],
                "generationConfig": {
                    "responseMimeType": "application/json",
                    "responseSchema": {
                        "type": "ARRAY",
                        "items": {
                            "type": "OBJECT",
                            "properties": {
                                "category": {"type": "STRING", "description": "Категория, например: Автоматические выключатели, Кабель, Щиты"},
                                "name": {"type": "STRING", "description": "Полное техническое наименование оборудования"},
                                "brand": {"type": "STRING", "description": "Производитель/Бренд (EKF, IEK, Schneider Electric и т.д.), если нет ставяй пустую строку"},
                                "article": {"type": "STRING", "description": "Заводской артикул, код заказа или шифр. Если нет - пустая строка"},
                                "unit": {"type": "STRING", "description": "Единица измерения (шт, м, компл)"},
                                "quantity": {"type": "STRING", "description": "Количество (числом или строкой)"}
                            },
                            "required": ["category", "name", "brand", "article", "unit", "quantity"]
                        }
                    }
                }
            }
            
            try:
                response = requests.post(gemini_url, json=payload, headers=headers, timeout=180)
                res_json = response.json()
                
                # Обработка стандартных ошибок Google (вкл. High Demand)
                if 'error' in res_json:
                    status.update(label="Ошибка ИИ-сервера", state="error")
                    st.error(f"Сервер Google отклонил запрос: {res_json['error'].get('message', '')}")
                    st.info("Если ошибка 'high demand' - это ограничения бесплатного ключа на объемные файлы.")
                    st.stop()
                
                # Обработка нестандартных ошибок ProxyAPI (Model not supported и т.д.)
                if 'detail' in res_json:
                    status.update(label="Ошибка ProxyAPI", state="error")
                    st.error(f"Шлюз ProxyAPI вернул ошибку: {res_json['detail']}")
                    st.stop()
                
                if 'candidates' not in res_json:
                    status.update(label="Неизвестный ответ", state="error")
                    st.error("В ответе нет данных (candidates). Выводим сырой ответ сервера для отладки:")
                    st.json(res_json)
                    st.stop()
                
                text_response = res_json['candidates'][0]['content']['parts'][0]['text'].strip()
                parsed_items = json.loads(text_response)
                st.write(f"📊 ИИ успешно нашел и структурировал позиций: {len(parsed_items)}")
                
            except Exception as e:
                status.update(label="Критическая ошибка разбора", state="error")
                st.error(f"Не удалось обработать ответ. Ошибка: {str(e)}")
                if 'res_json' in locals():
                    st.json(res_json)
                st.stop()
                
            # --- ШАГ 3: Цены и Ссылки ---
            status.update(label="Шаг 3: Подготовка поисковых запросов к ETM...", state="running")
            final_data = []
            for item in parsed_items:
                price = 430.00  # Базовая заглушка цены
                try:
                    qty_str = str(item.get('quantity', '1')).replace(',', '.').strip()
                    qty = float(''.join(c for c in qty_str if c.isdigit() or c == '.'))
                except:
                    qty = 1.0
                
                article = item.get('article', '').strip()
                brand = item.get('brand', '').strip()
                search_query = article if article else item.get('name', '')
                
                verification_flag = "Точный поиск" if article else "Проверить вручную (нет артикула)"
                
                final_data.append({
                    "Категория": item.get('category', 'Разное'),
                    "Наименование материала / оборудования": item.get('name', '—'),
                    "Бренд": brand if brand else "—",
                    "Артикул": article if article else "—",
                    "Ед. изм.": item.get('unit', 'шт.'),
                    "Кол-во": qty,
                    "Цена (руб.)": price,
                    "Сумма (руб.)": None, 
                    "Статус проверки": verification_flag,
                    "Ссылка на ETM": f"https://www.etm.ru/catalog?search={search_query}"
                })
                
            # --- ШАГ 4: Профессиональная сборка Excel через openpyxl ---
            status.update(label="Шаг 4: Генерация брендированного отчета Excel...", state="running")
            
            df = pd.DataFrame(final_data)
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Смета проекта')
                workbook = writer.book
                worksheet = writer.sheets['Смета проекта']
                
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                data_font = Font(name="Arial", size=10)
                flag_font_attention = Font(name="Arial", size=10, color="9C0006", italic=True)
                total_font = Font(name="Arial", size=11, bold=True)
                
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9')
                )
                
                for col_num in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for row_num in range(2, len(df) + 2):
                    worksheet.cell(row=row_num, column=8, value=f"=F{row_num}*G{row_num}")
                    worksheet.cell(row=row_num, column=7).number_format = '#,##0.00'
                    worksheet.cell(row=row_num, column=8).number_format = '#,##0.00'
                    
                    etm_url = df.iloc[row_num-2]["Ссылка на ETM"]
                    worksheet.cell(row=row_num, column=10).value = f'=HYPERLINK("{etm_url}", "Открыть в ETM")'
                    worksheet.cell(row=row_num, column=10).font = Font(name="Arial", size=10, color="0563C1", underline="single")
                    
                    for col_num in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.border = thin_border
                        if col_num != 10 and col_num != 9:
                            cell.font = data_font
                        
                        if col_num == 9 and "Проверить" in str(cell.value):
                            cell.font = flag_font_attention
                        elif col_num == 9:
                            cell.font = data_font

                total_row = len(df) + 3
                worksheet.cell(row=total_row, column=7, value="ИТОГО ПО СМЕТЕ:").font = total_font
                worksheet.cell(row=total_row, column=8, value=f"=SUM(H2:H{total_row-1})").font = total_font
                worksheet.cell(row=total_row, column=8).number_format = '#,##0.00'
                
                for col in worksheet.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        if cell.value:
                            val_str = str(cell.value)
                            if "HYPERLINK" in val_str:
                                val_str = "Открыть в ETM"
                            max_len = max(max_len, len(val_str))
                    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
            processed_data = output.getvalue()
            status.update(label="Расчет завершен! Премиальный Excel сформирован.", state="complete")
            
        st.download_button(
            label="📥 Скачать готовую смету Excel",
            data=processed_data,
            file_name="КОРПОРАТИВНАЯ_СМЕТА_ИИ.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
